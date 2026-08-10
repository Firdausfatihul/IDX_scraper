from __future__ import annotations

import json
import threading
from datetime import datetime, timedelta, timezone
from pathlib import Path
from types import SimpleNamespace
from zoneinfo import ZoneInfo

import openpyxl

from idx_digest.config import Settings
from idx_digest.db import Database
from idx_digest.disclosure_classifier import disclosure_class
from idx_digest.extractors import extract_xlsx
from idx_digest.idx_client import IDXClient
from idx_digest.incremental import company_input_fingerprint, promote_single_announcement
from idx_digest.network_watchdog import NetworkWatchdog
from idx_digest.routine_triage import RoutineEvidence, evaluate_routine_disclosure
from idx_digest.stock_master import StockMaster, StockMasterCache, parse_stock_master_payload


def _ann_summary(ticker: str, aid: str, when: str = "2026-08-07T10:00:00+07:00") -> dict:
    return {
        "ticker": ticker, "announcement_id": aid, "announced_at": when, "title": "Material update",
        "executive_summary": f"{ticker} material update", "category": "material",
        "material_facts": ["New project disclosed"], "financial_figures": [], "corporate_actions": [],
        "expansion_projects": ["Capacity project"], "management_or_control_changes": [],
        "capital_structure_events": [], "listing_or_regulatory_events": [], "dates_and_deadlines": [],
        "analytical_scenarios": [], "risks_or_uncertainties": [], "possible_investor_relevance": ["Monitor execution"],
        "source_files": [], "claim_sources": [], "limitations": [],
    }


def test_company_fingerprint_and_single_announcement_promotion(tmp_path: Path) -> None:
    start, end = "2026-07-06T00:00:00+07:00", "2026-08-07T23:59:00+07:00"
    record = {"announcement_id":"a1", "announced_at":"2026-08-07T10:00:00+07:00", "title":"Material update", "summary":_ann_summary("ANTM","a1"), "source_model":"m", "source_prompt_version":"a"}
    fp = company_input_fingerprint(ticker="ANTM", start_at=start, end_at=end, announcements=[record], model="m", prompt_version="c")
    db = Database(tmp_path / "db.sqlite3")
    promoted = promote_single_announcement(ticker="ANTM", start_at=start, end_at=end, record=record)
    assert promoted["announcement_count"] == 1
    assert promoted["overview"] == "ANTM material update"
    db.save_company_summary("ANTM", start, end, promoted, "m", "c", input_fingerprint=fp, generation_mode="single_announcement_promotion", source_announcement_count=1)
    assert db.company_summary_is_current("ANTM", start, end, model="m", prompt_version="c", input_fingerprint=fp)
    changed = dict(record); changed["summary"] = dict(record["summary"], executive_summary="changed")
    fp2 = company_input_fingerprint(ticker="ANTM", start_at=start, end_at=end, announcements=[changed], model="m", prompt_version="c")
    assert fp2 != fp
    assert not db.company_summary_is_current("ANTM", start, end, model="m", prompt_version="c", input_fingerprint=fp2)


def _item(aid: str, ticker: str = "ANTM") -> dict:
    return {"pengumuman":{"Id2":aid,"Kode_Emiten":ticker,"TglPengumuman":"2026-08-07T10:00:00","JudulPengumuman":"x"},"attachments":[]}


def test_pagination_inconsistency_falls_back_to_daily_shards(tmp_path: Path, monkeypatch) -> None:
    settings = Settings(data_dir=tmp_path / "data", idx_request_delay_seconds=0, idx_page_size=50, stock_master_enabled=False)
    client = IDXClient(settings)
    calls=[]
    def fake(params):
        calls.append(dict(params))
        if params["dateFrom"] != params["dateTo"]:
            if params["indexFrom"] == 0: return {"ResultCount":4374,"Replies":[_item(f"p{i}") for i in range(50)]}
            if params["indexFrom"] == 50: return {"ResultCount":4374,"Replies":[_item(f"q{i}") for i in range(50)]}
            return {"ResultCount":4374,"Replies":[]}
        aid = "day-" + params["dateFrom"]
        return {"ResultCount":1,"Replies":[_item(aid)]}
    monkeypatch.setattr(client, "_get_json", fake)
    tz=ZoneInfo("Asia/Jakarta")
    items, diag = client.collect_announcements(datetime(2026,8,6,tzinfo=tz), datetime(2026,8,7,23,59,tzinfo=tz))
    client.close()
    assert diag["complete"] is True
    assert diag["strategy"] == "date-shards"
    assert {x["pengumuman"]["Id2"] for x in items} == {"day-20260806","day-20260807"}
    assert any(c["indexFrom"] == 100 for c in calls)


def test_stock_master_parser_and_stale_cache(tmp_path: Path) -> None:
    assert parse_stock_master_payload({"Results":[{"KodeEmiten":"ANTM"},{"Ticker":"BBCA"}]}) == frozenset({"ANTM","BBCA"})
    path=tmp_path/"stock_master.json"; cache=StockMasterCache(path)
    old=(datetime.now(timezone.utc)-timedelta(hours=48)).isoformat()
    cache.save(StockMaster(frozenset({"ANTM"}),"test",old))
    assert cache.load() is not None
    assert cache.load(max_age_hours=24) is None


def test_public_expose_is_high_value_class() -> None:
    for title in ("Paparan Publik 2026", "Public Expose", "Penyampaian Materi pada Investor Meeting and Connectivity 2026", "Pemaparan Kinerja Perusahaan"):
        assert disclosure_class(title) == "public_expose"
    assert disclosure_class("Laporan Bulanan Registrasi Pemegang Efek") == "general"


def test_routine_delta_exposes_numeric_signal_details() -> None:
    decision = evaluate_routine_disclosure(
        "Laporan Bulanan Registrasi Pemegang Efek",
        [RoutineEvidence(filename="holder.pdf", text="Kepemilikan pemegang saham 5.00% menjadi 12.00%." * 8)],
        max_characters=70000, ownership_delta_threshold_pct=0.10,
    )
    assert decision.mode == "full"
    assert "ownership-percentage-delta" in decision.signals
    assert decision.signal_details
    detail = decision.signal_details[0]
    assert float(detail["absolute_delta_pct_points"]) >= 7.0


def test_network_watchdog_recovers_after_transport_failure(monkeypatch) -> None:
    events=[]
    observer=SimpleNamespace(event=lambda stage,message,**fields: events.append((stage,message,fields)))
    w=NetworkWatchdog("https://openrouter.ai/api/v1",observer=observer,probe_interval=.5,probe_timeout=.2)
    w.record_failure(TimeoutError("read operation timed out"))
    monkeypatch.setattr(w,"_probe",lambda: True)
    w.before_request()
    assert w.metrics["offline"] is False
    assert w.metrics["outage_count"] == 1
    assert w.metrics["recovery_count"] == 1
    assert any("restored" in message for _,message,_ in events)


def test_xlsx_primary_sheets_are_ranked_and_formula_fallback_is_visible(tmp_path: Path) -> None:
    path=tmp_path/"financial.xlsx"; wb=openpyxl.Workbook(); ws=wb.active; ws.title="Token"; ws["A1"]="taxonomy"; ws2=wb.create_sheet("Statement of Financial Position"); ws2["A1"]="Assets"; ws2["B1"]="=1+2"; wb.save(path)
    result=extract_xlsx(path,SimpleNamespace(max_xlsx_cells=1000),None)
    assert result.text.index("Statement of Financial Position") < result.text.index("===== SHEET: Token =====")
    assert "FORMULA_NO_CACHED_VALUE" in result.text


def test_same_window_rerun_reuses_company_fingerprint(tmp_path: Path, monkeypatch) -> None:
    import time
    import idx_digest.pipeline as pipeline_module
    from idx_digest.pipeline import Pipeline, PreparedAttachment

    def item(i: int):
        return {"pengumuman":{"Id2":f"a{i}","Kode_Emiten":"ANTM","TglPengumuman":f"2026-08-07T10:0{i}:00","JudulPengumuman":f"Update {i}","NoPengumuman":str(i),"JenisPengumuman":"General","PerihalPengumuman":"General"},"attachments":[{"FullSavePath":f"https://example.test/{i}.pdf","OriginalFilename":f"{i}.pdf","IsAttachment":False}]}
    class IDX:
        def __init__(self,*a,**k): self.browser=None
        def iter_announcements(self,*a,**k): yield item(1); yield item(2)
        def close(self): pass
        def browser_transport(self): raise AssertionError
    class Downloader:
        def __init__(self,*a,**k): pass
        def close(self): pass
    class S:
        model="m"; document_prompt_version="d"; announcement_prompt_version="a"; company_prompt_version="c"
        def __init__(self): self.prompts=SimpleNamespace(profile_name="p",hashes={}); self.company_calls=0
        @staticmethod
        def is_valid_document_summary(x): return bool(x and x.get("summary"))
        @staticmethod
        def is_valid_announcement_summary(x): return bool(x and x.get("executive_summary"))
        def summarize_document(self,*,ticker,filename,text,**kw): return {"ticker":ticker,"summary":text,"chunk_count":1}
        def summarize_announcement(self,*,announcement,documents,**kw): return _ann_summary("ANTM",announcement["id2"],announcement["announced_at"])
        def summarize_company_window(self,*,ticker,start_at,end_at,announcements,**kw):
            self.company_calls += 1
            return {"ticker":ticker,"period":{"start":start_at,"end":end_at},"announcement_count":len(announcements),"overview":"combined","timeline":[],"material_changes":[],"key_financial_figures":[],"corporate_actions":[],"expansion_projects":[],"management_or_control_changes":[],"capital_structure_events":[],"listing_or_regulatory_events":[],"analytical_scenarios":[],"risks_or_uncertainties":[],"items_to_monitor":[],"claim_sources":[],"limitations":[]}
        def close(self): pass
    monkeypatch.setattr(pipeline_module,"IDXClient",IDX); monkeypatch.setattr(pipeline_module,"AttachmentDownloader",Downloader)
    settings=Settings(data_dir=tmp_path/"data",openrouter_api_key="x",idx_request_delay_seconds=0,llm_concurrency=2,llm_per_announcement_concurrency=1,stock_master_enabled=False)
    fake=S()
    def build():
        p=Pipeline(settings,skip_llm=True); p.summarizer=fake
        def prepare(_d,*,announcement_id,ticker,attachment):
            path=tmp_path/f"{announcement_id}.txt"; path.write_text("evidence")
            return PreparedAttachment(url=attachment["FullSavePath"],filename=attachment["OriginalFilename"],text_path=path)
        monkeypatch.setattr(p,"_download_or_cached_attachment",prepare); monkeypatch.setattr(p,"_refresh_share_exports",lambda *a,**k:{}); monkeypatch.setattr(p,"_export_company_checkpoint",lambda *a,**k:None)
        return p
    tz=ZoneInfo("Asia/Jakarta"); start=datetime(2026,8,7,0,0,tzinfo=tz); end=datetime(2026,8,7,23,59,tzinfo=tz)
    first=build().run(start_at=start,end_at=end)
    assert fake.company_calls == 1 and first["company_cache"]["rebuilding"] == 1
    second=build().run(start_at=start,end_at=end)
    assert fake.company_calls == 1
    # v0.15.5 short-circuits only a fully covered incremental window before any
    # IDX metadata request, attachment work, or company-cache planning.
    assert second["metadata_noop"] is True
    assert second["metadata_cached_duplicates"] == 0
    assert second["company_cache"]["rebuilding"] == 0


def test_public_expose_prompt_has_separate_cache_lineage(tmp_path: Path) -> None:
    from idx_digest.prompts import PromptStore
    bundle=PromptStore(tmp_path/"prompts.json").load()
    assert "public_expose_document" in bundle.prompts
    general=bundle.document_version(schema_version="document-v1")
    public=bundle.public_expose_document_version(schema_version="document-v1")
    assert public != general
    text=bundle.prompts["public_expose_document"].lower()
    assert "guidance" in text and "capex" in text and "public expose" in text


def test_v015_does_not_reduce_llm_generation_ceilings() -> None:
    import inspect
    import idx_digest.summarizer as sm
    source=inspect.getsource(sm.OpenRouterSummarizer)
    for ceiling in ("max_tokens=6500", "max_tokens=7000", "max_tokens=5500", "max_tokens=10000"):
        assert ceiling in source


def test_v015_company_cache_columns_migrate_additively(tmp_path: Path) -> None:
    import sqlite3
    path=tmp_path/"old.sqlite3"
    with sqlite3.connect(path) as conn:
        conn.execute("CREATE TABLE company_window_summaries (ticker TEXT NOT NULL,start_at TEXT NOT NULL,end_at TEXT NOT NULL,summary_json TEXT NOT NULL,model TEXT NOT NULL,prompt_version TEXT NOT NULL,updated_at TEXT NOT NULL,PRIMARY KEY(ticker,start_at,end_at))")
        conn.execute("INSERT INTO company_window_summaries VALUES (?,?,?,?,?,?,?)",("ANTM","s","e",json.dumps({"ticker":"ANTM","overview":"legacy"}),"m","p","now"))
    db=Database(path)
    with db.connect() as conn:
        cols={r[1] for r in conn.execute("PRAGMA table_info(company_window_summaries)")}
        row=conn.execute("SELECT overview FROM (SELECT json_extract(summary_json,'$.overview') AS overview FROM company_window_summaries WHERE ticker='ANTM')").fetchone()
    assert {"input_fingerprint","generation_mode","source_announcement_count"} <= cols
    assert row[0] == "legacy"


def test_run_record_persists_progressive_company_summary(tmp_path: Path) -> None:
    from idx_digest.gui import RunRecord
    record=RunRecord(run_id="r1",request={},storage_dir=tmp_path/"r1")
    summary={"ticker":"ANTM","overview":"progressive"}
    record.publish({"type":"event","stage":"company-cache","message":"company summary cache hit","timestamp":"2026-08-09T09:00:00+07:00","fields":{"ticker":"ANTM","summary":summary}})
    assert record.snapshot()["summaries"]["ANTM"]["overview"] == "progressive"
    saved=json.loads((tmp_path/"r1"/"state.json").read_text())
    assert saved["summaries"]["ANTM"]["overview"] == "progressive"


def test_run_manager_marks_persisted_running_job_interrupted(tmp_path: Path) -> None:
    from idx_digest.gui import RunManager
    runs=tmp_path/"data"/"runs"; rd=runs/"deadbeef"; rd.mkdir(parents=True)
    (rd/"state.json").write_text(json.dumps({"run_id":"deadbeef","request":{"mode":"scrape","start":"2026-08-07T00:00","end":"2026-08-07T23:59"},"status":"running","created_at":"2026-08-09T08:00:00+07:00","last_seq":0}),encoding="utf-8")
    settings=Settings(data_dir=tmp_path/"data",openrouter_api_key="x")
    manager=RunManager(runs_root=runs,settings=settings)
    snap=manager.get("deadbeef").snapshot()
    assert snap["status"] == "interrupted"
