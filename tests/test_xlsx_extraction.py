from pathlib import Path
from types import SimpleNamespace

import openpyxl

from idx_digest.extractors import extract_xlsx


def test_xlsx_extractor_trims_blank_formatted_tail_and_labels_rows(tmp_path: Path):
    path = tmp_path / "financial.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Statement of Financial Position"
    ws["A1"] = "Assets"
    ws["B1"] = 1000
    # Create a formatted blank cell far to the right. It should not burn through
    # the extraction cell budget for every row.
    ws["ZZ1"].number_format = "#,##0"
    ws["A3"] = "Liabilities"
    ws["B3"] = 400
    wb.save(path)

    result = extract_xlsx(path, SimpleNamespace(max_xlsx_cells=20), None)
    assert "===== SHEET: Statement of Financial Position =====" in result.text
    assert "ROW 1 | Assets | 1000" in result.text
    assert "ROW 3 | Liabilities | 400" in result.text
    assert "TRUNCATED" not in result.text
